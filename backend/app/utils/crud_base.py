"""通用 CRUD 基类 - 提供标准增删改查、批量操作、排序、筛选、导入导出"""
import io
import csv
from typing import Any, Dict, Generic, List, Optional, Tuple, Type, TypeVar

from fastapi import HTTPException, Query, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import asc, desc, func, or_, Boolean
from sqlalchemy.orm import Session

from app.database import Base

ModelType = TypeVar("ModelType", bound=Base)
CreateSchemaType = TypeVar("CreateSchemaType", bound=BaseModel)
UpdateSchemaType = TypeVar("UpdateSchemaType", bound=BaseModel)


class CRUDBase(Generic[ModelType, CreateSchemaType, UpdateSchemaType]):
    """通用 CRUD 操作"""

    def __init__(self, model: Type[ModelType]):
        self.model = model

    def get(self, db: Session, id: int) -> Optional[ModelType]:
        return db.query(self.model).filter(self.model.id == id).first()

    def get_multi(
        self,
        db: Session,
        *,
        skip: int = 0,
        limit: int = 20,
        sort_field: Optional[str] = None,
        sort_order: str = "asc",
        filters: Optional[Dict[str, Any]] = None,
        search: Optional[str] = None,
        search_fields: Optional[List[str]] = None,
    ) -> Tuple[List[ModelType], int]:
        query = db.query(self.model)

        # 筛选
        if filters:
            for key, value in filters.items():
                if hasattr(self.model, key) and value is not None and value != "":
                    column = getattr(self.model, key)
                    if isinstance(value, str) and "%" in value:
                        query = query.filter(column.like(value))
                    else:
                        query = query.filter(column == value)

        # 搜索
        if search and search_fields:
            conditions = []
            for field in search_fields:
                if hasattr(self.model, field):
                    col = getattr(self.model, field)
                    if hasattr(col, "like"):
                        conditions.append(col.like(f"%{search}%"))
            if conditions:
                query = query.filter(or_(*conditions))

        # 统计总数
        total = query.count()

        # 排序
        if sort_field and hasattr(self.model, sort_field):
            sort_col = getattr(self.model, sort_field)
            if sort_order == "desc":
                query = query.order_by(desc(sort_col))
            else:
                query = query.order_by(asc(sort_col))
        elif hasattr(self.model, "sort_order"):
            query = query.order_by(asc(self.model.sort_order))
        else:
            query = query.order_by(desc(self.model.id))

        items = query.offset(skip).limit(limit).all()
        return items, total

    def create(self, db: Session, *, obj_in: CreateSchemaType) -> ModelType:
        obj_data = obj_in.model_dump()
        # 设置排序号
        if hasattr(self.model, "sort_order"):
            max_order = db.query(func.max(self.model.sort_order)).scalar() or 0
            obj_data["sort_order"] = max_order + 1
        db_obj = self.model(**obj_data)
        db.add(db_obj)
        db.commit()
        db.refresh(db_obj)
        return db_obj

    def create_with_dict(self, db: Session, *, obj_data: dict) -> ModelType:
        """使用预处理的字典创建记录（跳过 schema 字段映射）"""
        if hasattr(self.model, "sort_order"):
            max_order = db.query(func.max(self.model.sort_order)).scalar() or 0
            obj_data["sort_order"] = max_order + 1
        db_obj = self.model(**obj_data)
        db.add(db_obj)
        db.commit()
        db.refresh(db_obj)
        return db_obj

    def update(self, db: Session, *, id: int, obj_in: UpdateSchemaType) -> ModelType:
        db_obj = self.get(db, id=id)
        if not db_obj:
            raise HTTPException(status_code=404, detail=f"记录不存在 (id={id})")
        update_data = obj_in.model_dump(exclude_unset=True)
        for field, value in update_data.items():
            if value is not None and hasattr(db_obj, field):
                # 将 0 转为 None（外键字段清空时前端可能发 0）
                if value == 0 and field.endswith("_id"):
                    value = None
                setattr(db_obj, field, value)
        db.commit()
        db.refresh(db_obj)
        return db_obj

    def update_with_dict(self, db: Session, *, id: int, obj_data: dict) -> ModelType:
        """使用预处理的字典更新记录（跳过 schema 字段映射）"""
        db_obj = self.get(db, id=id)
        if not db_obj:
            raise HTTPException(status_code=404, detail=f"记录不存在 (id={id})")
        for field, value in obj_data.items():
            if value is not None and hasattr(db_obj, field):
                if value == 0 and field.endswith("_id"):
                    value = None
                setattr(db_obj, field, value)
        db.commit()
        db.refresh(db_obj)
        return db_obj

    def delete(self, db: Session, *, id: int) -> ModelType:
        db_obj = self.get(db, id=id)
        if not db_obj:
            raise HTTPException(status_code=404, detail=f"记录不存在 (id={id})")
        db.delete(db_obj)
        db.commit()
        return db_obj

    def batch_delete(self, db: Session, *, ids: List[int]) -> int:
        deleted = db.query(self.model).filter(self.model.id.in_(ids)).delete(synchronize_session=False)
        db.commit()
        return deleted

    def batch_update_status(self, db: Session, *, ids: List[int], status_field: str, value: str) -> int:
        if hasattr(self.model, status_field):
            count = (
                db.query(self.model)
                .filter(self.model.id.in_(ids))
                .update({status_field: value}, synchronize_session=False)
            )
            db.commit()
            return count
        return 0

    def update_sort_order(self, db: Session, *, id: int, sort_order: int) -> ModelType:
        db_obj = self.get(db, id=id)
        if not db_obj:
            raise HTTPException(status_code=404, detail=f"记录不存在 (id={id})")
        db_obj.sort_order = sort_order
        db.commit()
        db.refresh(db_obj)
        return db_obj

    def export_csv(self, db: Session, *, filters: Optional[Dict] = None) -> str:
        items, _ = self.get_multi(db, skip=0, limit=10000, filters=filters)
        if not items:
            return "无数据"
        output = io.StringIO()
        writer = csv.writer(output)
        # 获取非关系型字段名
        columns = [c.name for c in self.model.__table__.columns if c.name != "password_hash"]
        writer.writerow(columns)
        for item in items:
            row = []
            for col in columns:
                val = getattr(item, col, "")
                if val is None:
                    val = ""
                row.append(str(val))
            writer.writerow(row)
        return output.getvalue()

    def import_excel(self, db: Session, file: UploadFile) -> Dict[str, Any]:
        """从 Excel/CSV 导入数据（支持 .xlsx 和 .csv）"""
        filename = (file.filename or "").lower()

        # ── CSV 导入 ──
        if filename.endswith(".csv"):
            raw = file.file.read()
            # 自动检测编码（尝试常见中文编码）
            encoding = "utf-8"
            for enc in ["utf-8-sig", "gbk", "gb2312", "gb18030", "utf-8"]:
                try:
                    raw.decode(enc)
                    encoding = enc
                    break
                except (UnicodeDecodeError, LookupError):
                    continue
            content = raw.decode(encoding, errors="replace")
            reader = csv.DictReader(io.StringIO(content))
            headers = [h.strip() for h in reader.fieldnames] if reader.fieldnames else []
            rows = list(reader)
            if len(rows) < 1:
                return {"imported": 0, "errors": ["数据不足"]}
            imported = 0
            errors = []
            for row_idx, row in enumerate(rows, start=2):
                try:
                    data = {}
                    for h in headers:
                        val = row.get(h, "").strip()
                        if val != "":
                            data[h] = val
                    if not data:
                        continue
                    # 跳过 id（自增）
                    data.pop("id", None)
                    data.pop("created_at", None)
                    data.pop("updated_at", None)
                    # 处理布尔字段（CSV 中的 TRUE/FALSE 字符串）
                    for col in self.model.__table__.columns:
                        if col.name in data and isinstance(col.type, Boolean):
                            val = data[col.name]
                            if isinstance(val, str):
                                data[col.name] = val.upper() in ("TRUE", "1", "YES", "是")
                    # 处理密码哈希
                    if hasattr(self.model, "password_hash") and "password_hash" not in data:
                        import hashlib
                        data["password_hash"] = hashlib.md5(("pwd_147258").encode()).hexdigest()
                    # 检查唯一约束字段（如 phone）是否已存在
                    skip = False
                    for col in self.model.__table__.columns:
                        if col.unique and col.name in data:
                            existing = db.query(self.model).filter(
                                getattr(self.model, col.name) == data[col.name]
                            ).first()
                            if existing:
                                errors.append(f"第{row_idx}行: {col.name}「{data[col.name]}」已存在，跳过")
                                skip = True
                                break
                    if skip:
                        continue
                    db_obj = self.model(**data)
                    if hasattr(self.model, "sort_order"):
                        max_order = db.query(func.max(self.model.sort_order)).scalar() or 0
                        db_obj.sort_order = max_order + 1
                    db.add(db_obj)
                    imported += 1
                except Exception as e:
                    errors.append(f"第{row_idx}行: {str(e)}")
                    continue
            db.commit()
            return {"imported": imported, "errors": errors[:10]}

        # ── XLSX 导入（原有逻辑） ──
        wb = load_workbook(file.file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {"imported": 0, "errors": ["数据不足"]}

        headers = [str(h).strip() for h in rows[0] if h]
        imported = 0
        errors = []

        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                data = {}
                for col_idx, header in enumerate(headers):
                    if col_idx < len(row) and row[col_idx] is not None:
                        data[header] = row[col_idx]
                if data:
                    # 处理密码哈希（导入数据不含此字段）
                    if hasattr(self.model, "password_hash") and "password_hash" not in data:
                        import hashlib
                        data["password_hash"] = hashlib.md5(("pwd_147258").encode()).hexdigest()
                    # 检查唯一约束字段是否已存在
                    skip = False
                    for col in self.model.__table__.columns:
                        if col.unique and col.name in data:
                            existing = db.query(self.model).filter(
                                getattr(self.model, col.name) == data[col.name]
                            ).first()
                            if existing:
                                errors.append(f"第{row_idx}行: {col.name}「{data[col.name]}」已存在，跳过")
                                skip = True
                                break
                    if skip:
                        continue
                    # 使用 create 方法
                    db_obj = self.model(**data)
                    if hasattr(self.model, "sort_order"):
                        max_order = db.query(func.max(self.model.sort_order)).scalar() or 0
                        db_obj.sort_order = max_order + 1
                    db.add(db_obj)
                    imported += 1
            except Exception as e:
                errors.append(f"第{row_idx}行: {str(e)}")
                continue

        db.commit()
        wb.close()
        return {"imported": imported, "errors": errors[:10]}  # 最多返回10个错误